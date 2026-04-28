import json
import pandas as pd
import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

STATUS_CODE_CANDIDATES = ['status', '状态码', 'status_code', 'code', 'Status', 'STATUS', 'J']
URL_CANDIDATES = ['url', 'URL', '网址', '链接', 'directurl', 'direct_url', 'Direct URL', 'E']
SURVIVALSCAN_COLUMNS = ['URL', 'Status', 'Status Code', 'Title', 'Alive', 'Source']

# 保持原始字段名，避免与后续流程的列识别逻辑不一致
COLUMN_MAPPING = {}


def extract_names(data):
    """从嵌套字典中提取所有'name'字段并拼接"""
    names = []
    try:
        if isinstance(data, str):
            data = json.loads(data)
        for key, value in data.items():
            if isinstance(value, dict) and 'name' in value:
                names.append(value['name'])
    except (json.JSONDecodeError, AttributeError, TypeError):
        return ""
    return ' | '.join(names)


def normalize_column_name(value):
    return str(value).strip().lower()



def find_column(columns, candidates):
    normalized_map = {}
    for column in columns:
        normalized_map.setdefault(normalize_column_name(column), column)

    for candidate in candidates:
        matched = normalized_map.get(normalize_column_name(candidate))
        if matched is not None:
            return matched
    return None



def extract_urls(df):
    url_col = find_column(df.columns, URL_CANDIDATES)
    if url_col is None:
        return [], None

    urls = df[url_col].dropna().astype(str).str.strip()
    urls = urls[urls != ''].drop_duplicates().tolist()
    return urls, url_col



def count_status_200(df):
    status_col = find_column(df.columns, STATUS_CODE_CANDIDATES)
    if status_col is None:
        return None, 0

    status_series = pd.to_numeric(df[status_col], errors='coerce')
    return status_col, int((status_series == 200).sum())



def detect_json_input_kind(input_file):
    with open(input_file, 'r', encoding='utf-8') as f:
        while True:
            char = f.read(1)
            if not char:
                return 'empty'
            if not char.isspace():
                return 'survivalscan_report' if char == '[' else 'spray_jsonl'



def load_survivalscan_report(input_file):
    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError('Web-SurvivalScan report.json 不是数组格式')
    return data



def normalize_survivalscan_status(status_value):
    return str(status_value).strip().lower()



def translate_survivalscan_status(status_value):
    status = normalize_survivalscan_status(status_value)
    if status == 'servival':
        return '存活'
    if status == 'deaed':
        return '失败'
    if status == 'reject':
        return '拒绝'
    return str(status_value).strip()



def parse_status_code(value):
    try:
        if value in ('', None):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None



def normalize_survivalscan_to_compat_df(report_data):
    rows = []
    for item in report_data:
        if not isinstance(item, dict):
            continue

        url = str(item.get('url', '')).strip()
        if not url:
            continue

        title_value = item.get('title', '')
        title = '' if title_value is None else str(title_value).strip()
        raw_status = item.get('status', '')
        normalized_status = normalize_survivalscan_status(raw_status)
        status_code = parse_status_code(item.get('statusCode'))
        is_alive = normalized_status == 'servival'

        if not is_alive:
            continue

        rows.append({
            'URL': url,
            'Status': translate_survivalscan_status(raw_status),
            'Status Code': status_code,
            'Title': title,
            'Alive': '是',
            'Source': 'Web-SurvivalScan',
        })

    return pd.DataFrame(rows, columns=SURVIVALSCAN_COLUMNS)



def beautify_spray_excel(file_path):
    """美化spray生成的Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        data_font = Font(color="000000", size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        headers = [cell.value for cell in ws[1]]
        status_header = find_column(headers, STATUS_CODE_CANDIDATES)
        if status_header is not None:
            col_idx = headers.index(status_header) + 1
            status_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
            ws.conditional_formatting.add(
                status_range,
                ColorScaleRule(
                    start_type='min',
                    start_color='FFC7CE',
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFFFCC',
                    end_type='max',
                    end_color='C6EFCE'
                )
            )

        wb.save(file_path)
        print(f"Spray Excel表格美化完成: {file_path}")
    except Exception as e:
        print(f"美化Spray Excel失败: {e}")



def beautify_compat_excel(file_path):
    """美化由 Web-SurvivalScan 生成的兼容结果表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "存活结果"
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        data_font = Font(color="000000", size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 60)

        headers = [cell.value for cell in ws[1]]
        url_header = find_column(headers, ['URL', 'url', '网址'])
        if url_header is not None:
            col_idx = headers.index(url_header) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith(('http://', 'https://')):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

        status_code_header = find_column(headers, ['Status Code', '状态码'])
        if status_code_header is not None and ws.max_row > 1:
            col_idx = headers.index(status_code_header) + 1
            status_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
            ws.conditional_formatting.add(
                status_range,
                CellIsRule(operator='equal', formula=['200'], fill=PatternFill(fill_type='solid', start_color='C6EFCE', end_color='C6EFCE'))
            )
            ws.conditional_formatting.add(
                status_range,
                CellIsRule(operator='equal', formula=['403'], fill=PatternFill(fill_type='solid', start_color='FFEB9C', end_color='FFEB9C'))
            )

        summary_ws = wb.create_sheet(title="汇总信息")
        summary_ws['A1'] = "Web 存活结果汇总"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A3'] = "总记录数:"
        summary_ws['B3'] = max(ws.max_row - 1, 0)
        summary_ws['A4'] = "来源:"
        summary_ws['B4'] = "Web-SurvivalScan"

        status_code_header = find_column(headers, ['Status Code', '状态码'])
        if status_code_header is not None:
            col_idx = headers.index(status_code_header) + 1
            status_codes = [ws.cell(row=row, column=col_idx).value for row in range(2, ws.max_row + 1)]
            status_codes = pd.to_numeric(pd.Series(status_codes), errors='coerce')
            summary_ws['A6'] = 'HTTP 200 数量:'
            summary_ws['B6'] = int((status_codes == 200).sum())
            summary_ws['A7'] = 'HTTP 403 数量:'
            summary_ws['B7'] = int((status_codes == 403).sum())

        wb.save(file_path)
        print(f"兼容结果表格美化完成: {file_path}")
    except Exception as e:
        print(f"美化兼容结果表失败: {e}")



def beautify_ehole_excel(file_path):
    """深度美化ehole生成的Excel表格"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        if ws.max_row <= 1:
            print(f"警告: ehole结果表格为空: {file_path}")
            return

        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        data_font = Font(color="000000", size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 60)

        url_cols = ['URL', 'url', '网址']
        for col_name in url_cols:
            if col_name in [cell.value for cell in ws[1]]:
                col_idx = [cell.value for cell in ws[1]].index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith(('http', 'https')):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")

        risk_cols = ['Risk', '风险等级', '危险程度']
        for col_name in risk_cols:
            if col_name in [cell.value for cell in ws[1]]:
                col_idx = [cell.value for cell in ws[1]].index(col_name) + 1
                risk_range = f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}"
                ws.conditional_formatting.add(
                    risk_range,
                    CellIsRule(operator='containsText', formula=['"高"'], fill=PatternFill(bgColor='FFC7CE'), font=Font(color='9C0006'))
                )
                ws.conditional_formatting.add(
                    risk_range,
                    CellIsRule(operator='containsText', formula=['"中"'], fill=PatternFill(bgColor='FFEB9C'), font=Font(color='9C5700'))
                )
                ws.conditional_formatting.add(
                    risk_range,
                    CellIsRule(operator='containsText', formula=['"低"'], fill=PatternFill(bgColor='C6EFCE'), font=Font(color='006100'))
                )
                break

        if ws.max_row > 10:
            try:
                data_range = f"'{ws.title}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
                pivot_ws = wb.create_sheet(title="数据透视表")
                pivot_ws['A1'] = "指纹识别结果统计"
                pivot_ws['A1'].font = Font(bold=True, size=16)

                from openpyxl.pivot.table import PivotTable

                pt = PivotTable(srcRange=data_range, dest=f"'{pivot_ws.title}'!$A$3", name="指纹统计")
                pt.addRow('A')
                if ws.max_column >= 2:
                    pt.addColumn('B')
                pt.addData('A', function='count')
                pivot_ws.add_pivot(pt)

                for row in pivot_ws.iter_rows(min_row=3, max_row=3):
                    for cell in row:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")

                if ws.max_column >= 3:
                    filter_values = list({ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)})
                    dv = DataValidation(type="list", formula1='"{}"'.format(','.join([str(v) for v in filter_values if v])))
                    dv.add(pivot_ws['D1'])
                    pivot_ws.add_data_validation(dv)
                    pivot_ws['C1'] = "筛选:"
                    pivot_ws['C1'].font = Font(bold=True)

                print("已为ehole结果添加数据透视表")
            except Exception as e:
                print(f"创建数据透视表失败: {e}")

        summary_ws = wb.create_sheet(title="汇总信息")
        summary_ws['A1'] = "指纹识别结果汇总"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A3'] = "总记录数:"
        summary_ws['B3'] = ws.max_row - 1

        wb.save(file_path)
        print(f"Ehole Excel表格深度美化完成: {file_path}")
    except Exception as e:
        print(f"美化Ehole Excel失败: {e}")



def process_spray_json(input_file, output_file):
    print(f"开始处理spray结果: {input_file}")

    data_list = []
    with open(input_file, 'r', encoding='utf-8') as f:
        for line in f:
            try:
                data = json.loads(line.strip())
                data_list.append(data)
            except json.JSONDecodeError:
                print(f"警告: 无法解析JSON行: {line[:50]}...")

    if not data_list:
        print(f"错误: 文件 {input_file} 中没有有效JSON数据")
        return False

    df = pd.DataFrame(data_list)
    df = df.rename(columns=COLUMN_MAPPING)

    if 'redirect_url' in df.columns:
        print("检测到'redirect_url'列，已删除")
        df = df.drop(columns=['redirect_url'])

    for nested_col in ['plugins', 'extracts', 'finger', 'O']:
        if nested_col in df.columns:
            df[nested_col] = df[nested_col].apply(extract_names)

    status_col, status_200_count = count_status_200(df)
    if status_col is not None:
        print(f"检测到状态码列: {status_col}，其中 200 状态码数量: {status_200_count}")
    else:
        print("未检测到状态码列，将保留全部记录")

    df.to_excel(output_file, index=False)
    print(f"Excel文件已保存: {output_file}")

    txt_output = os.path.splitext(output_file)[0] + ".txt"
    urls, url_col = extract_urls(df)
    with open(txt_output, 'w', encoding='utf-8') as f:
        if urls:
            f.write('\n'.join(urls))

    if url_col is not None:
        print(f"已从列 {url_col} 提取 {len(urls)} 个URL保存到: {txt_output}")
    else:
        print(f"警告: 未找到URL列，已生成空URL文件: {txt_output}")

    beautify_spray_excel(output_file)
    return True



def process_survivalscan_report(input_file, output_file):
    print(f"开始处理 Web-SurvivalScan 报告: {input_file}")
    report_data = load_survivalscan_report(input_file)
    df = normalize_survivalscan_to_compat_df(report_data)
    df.to_excel(output_file, index=False)
    print(f"兼容Excel文件已保存: {output_file}")
    print(f"已保留 {len(df)} 条存活记录")
    beautify_compat_excel(output_file)
    return True



def process_data(input_file, output_file):
    """处理JSON输入文件，生成Excel和TXT输出"""
    try:
        file_ext = os.path.splitext(input_file)[1].lower()

        if file_ext == '.json':
            json_kind = detect_json_input_kind(input_file)
            if json_kind == 'empty':
                print(f"错误: 文件 {input_file} 为空")
                return
            if json_kind == 'survivalscan_report':
                process_survivalscan_report(input_file, output_file)
            else:
                process_spray_json(input_file, output_file)

        elif file_ext in ['.xlsx', '.xls']:
            print(f"开始美化ehole结果: {input_file}")
            if input_file != output_file:
                shutil.copy2(input_file, output_file)
            beautify_ehole_excel(output_file)

        else:
            print(f"错误: 不支持的文件类型: {file_ext}")
            sys.exit(1)

    except Exception as e:
        print(f"处理文件时出错: {e}")
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("用法: python process_data.py <输入JSON/Excel文件> <输出Excel文件>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    process_data(input_file, output_file)
    sys.exit(0)
